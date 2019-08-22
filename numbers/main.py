from openpyxl import load_workbook, Workbook
import csv

MAX_COLUMNS = 3
XLSX_FILE = 'newfile.csv'
# CREATE EMPTY newfile.csv in root folder near main.py
# THEN UPLOAD IT TO GOOGLE SPREADSHEES

def append_row(row):
    with open(XLSX_FILE, 'a') as csvFile:
        writer = csv.writer(csvFile)
        writer.writerow(row)


wb = load_workbook('base.xlsx')
ws = wb.active

newfile = Workbook()
wsn = newfile.active

# iterate over all rows
# for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
for row in ws.iter_rows(min_row=1, max_col=MAX_COLUMNS):
    number = str(row[2].value)
    converted = number
    if number:
        if number[0] == '(':
            print(number[1:])
            converted = number[1:]
        elif number[0] == '0':
            if len(number) == 10:
                print('+38' + number)
                converted = '+38' + number
            else:
                print(number)
                converted = number
        elif number[0] != '0':
            if len(number) == 9:
                print('+380' + number)
                converted = '+380' + number
            else:
                print(number)
                converted = number
    else:
        print(number)
        converted = number

    newrow = [row[0].value, row[1].value, converted]
    append_row(newrow)

print('\n\n')
print('#' * 50)
print('FINISHED')
